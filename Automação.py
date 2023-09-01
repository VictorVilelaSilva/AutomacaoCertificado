from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import shutil
from tkinter import N
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import re
import unidecode
import string
from difflib import SequenceMatcher
from nltk.corpus import stopwords
import nltk
import xlsxwriter
from functools import cache

stopwords = nltk.corpus.stopwords.words('portuguese')


@cache
def limpar_string(text):
    text = ''.join([word for word in text if word not in string.punctuation])
    text = text.lower()
    text = ' '.join([word for word in text.split() if word not in stopwords])
    text = re.sub(r"\s+", "", text)
    text = unidecode.unidecode(text)

    return text

# limpa string para formato do arquivo TEX


def limpar_string2(text):
    text = text.title()
    text = ''.join([word for word in text if word not in string.punctuation])
    text = ' '.join([word for word in text.split() if word not in stopwords])
    text = re.sub(r"\s+", "", text)
    text = unidecode.unidecode(text)

    return text


def similar(a, b):

    return SequenceMatcher(None, limpar_string(a), limpar_string(b)).ratio()


def trocarNome(search_text, replace_text, nome):
    with open(nome, 'r', encoding='utf-8') as file:
        data = file.read()
        data = data.replace(search_text, replace_text)
    with open(nome, 'w', encoding='utf-8') as file:
        file.write(data)


def tabela(arquivoNome, cursos, NomeProcurado):
    Atividades_Mocitec = {'Minicurso: A Arte do Origami': 2, 'Minicurso: Power Bi: Introdução à Análise de dados': 4, 'Oficina: Introdução à Linguística': 2,
                          'Minicurso: Python': 4, 'Minicurso: Introdução ao CLP': 4, 'Minicurso: Arduino com Python': 4, 'Oficina: Mitologia Grega': 2,
                          'Minicurso: Introdução à Modelagem 3D Utilizando Software Inventor': 2, 'Minicurso: Visão Computacional': 4, 'Minicurso: Introdução à Programação Competitiva': 4,
                          'Minicruso: Introdução ao LaTeX': 2, 'Minicurso: Planilhas Eletrônicas': 8, 'Minicurso: Colagem Digital': 2, 'Minicurso: SQL (Standard Query Language)': 4,
                          'Minicurso: Unity 2D: Jogo de Nave': 4, 'Minicurso: Arduino': 5, 'Minicurso: GIT': 2, 'Oficina: Preparação para Literatura PISM 1': 2, 'Oficina: Arte na Madeira': 4, 'Oficina: Teatro': 2, 'Minicurso: Introdução à Fresadora CNC': 4,
                          'Minicurso: Programação Competitiva': 4, 'Oficina: Spoken English Fluency Workshop': 6, 'Oficina: Xadrez': 4, 'Oficina: Show de Química': 2, 'Competições Acadêmicas: Competição de Combate Sumô LEGO': 4, 'Palestra: Tecnologia e Inovação - O que o mercado espera?': 2, 'Palestra: Desafios da Mulher na Sociedade Contemporânea': 2,
                          'Palestra: Soft Skills no Desenvolvimento de Software': 2, 'Palestra: Sistema de Distribuição de Energia Elétrica no Brasil': 2, 'Palestra: Security on Shopfloor - Visão Organizacional': 2, 'Palestra: Experiências no Trabalho Remoto em Desenvolvimento de Software na Healthtech iClinic': 2, 'Palestra: Da Ciência para a Inovação': 2, 'Palestra: Conhecendo um Pouco Sobre o Desenvolvimento do GloboPlay': 2, 'Visita Técnica:RBM Web - Leopoldina': 1,
                          'Visita Técnica: Racional Elétrica Serviço de Engenharia e Usina Solar em Ubá': 5, 'Visita Técnica: Embrapa gado de Leite - Campo Experimental': 7, 'Oficina: Capoeira': 2, 'Minicurso: Introdução à Inteligência Computacional Usando Google Colab': 12, 'Oficina: Matemática Básica - Turma 2': 2, 'Minicurso: 3D: Desenho e Impressão': 8, 'Equipe de Apoio: Organização das cadeiras na quadra': 2,
                          'Palestra: Mesa Redonda - RBM': 5, 'Equipe de Apoio: Interclasse': 2, 'Cerimônia de Abertura': 1, 'Cerimônia de Encerramento': 2, 'Oficina: Análise Fílmica - Narrativa e Representação no Cinema': 8, 'Competição de E-sports: League Of Legends': 3, 'Competição de E-sports: FIFA 22': 3, 'Minicurso: Sistema de Refrigeração e Instalação de Ar Condicionado SPLIT': 15,
                          'FalaÊ': 2, 'Oficina: Matemática Básica - Turma 1': 2, 'Oficina: Resistência dos Materiais - Estudo Dirigido': 4, 'Competições Acadêmcas: Competição de Programação': 3, 'Oficina: Mecânica Técnica - Estudo Dirigido': 4, 'Visita Técnica: UHE Barra do Braúna': 5, 'Concurso de Trabalhos': 6, 'Almoço Musical': 3,
                          'Ministrante: GIT': 2, 'Ministrante: Arduino com Python': 4, 'Ministrante: Introdução à Fresadora CNC': 4, 'Ministrante: Introdução à Modelagem 3D Utilizando Software Inventor': 2, 'Ministrante: Análise Fílmica - Narrativa e Representação no Cinema': 8, 'Ministrante: Arduino': 5, 'Ministrante: Arte na Madeira': 4,
                          'Ministrante: Planilhas Eletrônicas': 8, 'Ministrante: Introdução à Inteligência Computacional Usando Google Colab': 12, 'Ministrante: Introdução à Linguística': 2, 'Ministrante: Introdução à Programação Competitiva': 4, 'Ministrante: Introdução ao CLP': 4, 'Ministrante: Introdução ao LaTeX': 2, 'Ministrante: A Arte do Origami': 2,
                          'Ministrante: Matemática Básica - Turma 1': 2, 'Ministrante: Matemática Básica - Turma 2': 2, 'Ministrante: Python': 4, 'Ministrante: Unity 2D: Jogo de Nave': 4, 'Ministrante: SQL (Standard Query Language)': 4, 'Ministrante: Mitologia Grega': 2, 'Ministrante: Capoeira': 2, 'Ministrante: Power Bi: Introdução à Análise de dados': 4, 'Ministrante: Preparação para Literatura PISM 1': 2, 'Ministrante: Programação Competitiva': 4, 'Ministrante: Show de Química': 2, 'Ministrante: Spoken English Fluency Workshop': 6, 'Ministrante: Teatro': 2, 'Ministrante: Visão Computacional': 4, 'Ministrante: Xadrez': 4, 'Ministrante: Colagem Digital': 2,
                          'Ação Social: Por um dia das crianças mais feliz - Equipe: CAUT CONOSCO': 15, 'Ação Social: Por um dia das crianças mais feliz - Equipe: Galera da Peteca': 10, 'Ação Social: Por um dia das crianças mais feliz - Equipe: Controle da Computação': 8, 'Ação Social: Por um dia das crianças mais feliz': 4, 'Ação Social: Gincana SangueBom - Equipe: Doadores': 20, 'Ação Social: Gincana SangueBom - Equipe: Encautech': 15, 'Ação Social: Gincana SangueBom - Equipe: Sangue latino': 10, 'Ação Social: Gincana SangueBom': 5,
                          'Staff - Minicurso: Power BI - Introdução à Análise de Dados': 4, 'Staff - Minicurso: SQL (Standard Query Language)': 4, 'Staff - Minicurso: Programação Competitiva': 4, 'Staff - Oficina: Capoeira': 2, 'Staff - Almoço Musical - 13/09': 3, 'Staff - Exposições Acadêmicas: Concurso de Trabalhos': 6, 'Staff - Competição de E-Sports: League of Legends - 14/09': 2, 'Staff - Competição de E-Sports: League of Legends - 12/09': 2,
                          'Staff - Competição de E-Sports: League of Legends - 13/09': 2, 'Staff - Competição de E-Sports: League of Legends - 15/09': 2, 'Staff - Oficina: Xadrez - 15/09': 2, 'Staff - Oficina: Xadrez - 16/09': 2, 'Staff - Oficina: Análise Fílmica - Narrativa e Representação no Cinema': 8, 'Staff - Oficina: Matemática Básica - Turma 2': 2, 'Staff - Oficina: Spoken English Fluency Workshop': 6, 'Staff - Oficina: Show de Química': 2, 'Staff - Minicurso: Python': 4, 'Staff - Minicurso: Visão Computacional': 4, 'Staff - Minicurso: Unity2D - Jogo de Nave': 4, 'Staff - Palestra: Conhecendo um Pouco Sobre o Desenvolvimento do GloboPlay': 2, 'Staff - Minicurso: Introdução ao CLP': 4,
                          'Staff - Minicurso: Introdução à Inteligência Computacional Usando Google Colab - 15/09': 4, 'Staff – Minicurso: Introdução à Inteligência Computacional Usando Google Colab - 12/09': 4, 'Staff - Minicurso: Introdução à Inteligência Computacional Usando Google Colab - 13/09': 4, 'Staff - Palestra: Experiências no Trabalho Remoto em Desenvolvimento de Software na Healthtech iClinic': 2, 'Staff - Almoço Musical - 15/09': 2, 'Staff - Almoço Musical - 14/09': 2, 'Staff - Oficina: Introdução à Linguística': 2, 'Staff - Minicurso: GIT': 2, 'Staff - Palestra: Soft Skills no Desenvolvimento de Software': 2, 'Staff - Minicurso: Introdução à Fresadora CNC': 4,
                          'Staff – Minicurso: Arduino com Python': 4, 'Staff - Palestra: Security on Shopfloor': 2, 'Staff - Minicurso: Arduino': 5, 'Participou da organização do evento totalizando 20 horas': 20, 'Staff - Minicurso: Introdução à Modelagem 3D Usando o Software Inventor': 2, 'Staff - Oficina: Matemática Básica - Turma 1': 2, 'Staff - Competições Acadêmicas: Competição de Combate Sumô LEGO': 4, 'Staff - Minicurso: Introdução à Programação Competitiva': 4, 'Staff – Minicurso: Planilhas Eletrônicas - 12/09': 4, 'Staff – Minicurso: Planilhas Eletrônicas – 13/09': 4, 'Staff - Palestra: Os Desafios das Mulheres na Contemporaneidade': 2, 'Staff - Minicurso: Introdução ao LaTeX': 2, 'Staff - Oficina: Mitologia Grega': 2,
                          'Staff - Oficina: Arte na Madeira': 4,'Participou da organização do evento totalizando 20 horas':20,'Ministrante: Competições Acadêmicas: Competição de Combate Sumô LEGO':4, 'Registro audiovisual do evento':20, 'Staff – Palestra: Sistema de Distribuição de Energia Elétrica no Brasil':2,
                          'Futsal Feminino – MOCITEC 2022':2,'Futsal Masculino – MOCITEC 2022':2,'Vôlei – MOCITEC 2022':2}
    Pessoas20Horas = []
    total_Horas = 0
    for i in range(1, len(cursos)): 
        Atividades_Mocitec[cursos[i]]
        total_Horas += Atividades_Mocitec[cursos[i]]

        texto = str(cursos[i])+'&' + \
            str(Atividades_Mocitec[cursos[i]]) + r'\\ \hline'
        file = open(dest, 'r', encoding='utf-8')
        lines = file.readlines()
        file.close()

        lines.insert(49, texto + "\n")

        file = open(dest, 'w', encoding='utf-8')
        file.writelines(lines)
        file.close()
    trocarNome('TotalHoras', str(total_Horas), dest)

    if total_Horas >= 20:
        # Pessoas20Horas.append(NomeProcurado)
        return NomeProcurado
    return 0


def Nomes_Iguais(Pagina1):
    planilhaNomesMocitec = load_workbook('InscritosNoEvento')
# abri o txt
    pagina2 = planilhaNomesMocitec.active
# Nomes errados
    nomes1 = []
# Nomes certos
    nomes2 = []
    Nomes_Iguais1 = []
    curso1 = []


# Percorre as coluna

    for row in range(2, 3504):
        # pega o conteudo da celula da planilha e guarda em nomesTemp
        nomes_Errados = Pagina1['A'+str(row)].value
        nomes_Errados = str(nomes_Errados)

        if nomes_Errados[len(nomes_Errados)-1] == ' ':
            nomes_Errados = nomes_Errados[:-1]

        curso1.append(Pagina1['B'+str(row)].value)
        nomes1.append(nomes_Errados)

    for row in range(2, 492):
        Nomes_Certos = pagina2['D'+str(row)].value
        Nomes_Certos = str(Nomes_Certos)
        if Nomes_Certos[len(Nomes_Certos)-1] == ' ':
            Nomes_Certos = Nomes_Certos[:-1]

        nomes2.append(Nomes_Certos)

    print("tamanho nomes errados " + str(len(nomes1)))
    print("tamanho nomes certos " + str(len(nomes2)))
    print("tamanho cursos fun " + str(len(curso1)))
    taxa_maior = 0
    for NumeroNome1 in range(len(nomes1)):

        for NumeroNome2 in range(len(nomes2)):
            Nome1Limpo = limpar_string(nomes1[NumeroNome1])
            Nome2Limpo = limpar_string(nomes2[NumeroNome2])

            taxa_semelhança = similar(Nome1Limpo, Nome2Limpo)

            if taxa_semelhança >= 0.75:
                Nomes_Iguais1.append(nomes2[NumeroNome2])
                nomes2[NumeroNome2] = '-'

    print("tamanho nomes iguais " + str(len(Nomes_Iguais1)))
    return Nomes_Iguais1, curso1, nomes1


###############################################################################

# abri a planilha desejada
planilhaNomesCursos = load_workbook('PresencasNosCursos.xlsx')
# abri o txt
Pagina1 = planilhaNomesCursos.active
nomes = []
curso = []
Presencas = []
CursosPlanilha = []
Nomes_Iguais2 = []
Nomes_Iguais2, curso, nomes = Nomes_Iguais(Pagina1)
print("tamanho cursos " + str(len(curso)))
print("tamanho nome " + str(len(nomes)))
taxa_semelhança = 0
taxa_maior = 0

print(len(Nomes_Iguais2))

# print(Nomes_Iguais2)
for i in range(len(Nomes_Iguais2)):

    NomeProcurado = Nomes_Iguais2[i]
    NomeProcurado = NomeProcurado.lower()
    NomeProcurado = NomeProcurado.title()
    CursosPlanilha.append(NomeProcurado)
    #print('NomeProcurado '+str(NomeProcurado))
    for j in range(len(nomes)):
        if(limpar_string(NomeProcurado) == limpar_string(nomes[j])):
            CursosPlanilha.append(curso[j])
            nomes[j] = '-'

    f2 = open(r'Certificados\\' +
              str(limpar_string2(NomeProcurado)) + '.tex', 'x')
    f2 = open(r'Certificados\\' + str(limpar_string2(NomeProcurado)) +
              '.tex', 'w')
    src = 'Certificado_Modelo_Novo.txt'
    dest = (r'Certificados\\' + str(limpar_string2(NomeProcurado)) + ".tex")

    shutil.copy2(src, dest)
    trocarNome('NomeParticipante', str(NomeProcurado.title()), dest)
    testeRetorno =tabela(dest, CursosPlanilha, NomeProcurado)
    if testeRetorno != 0:
        Presencas.append(testeRetorno)

    CursosPlanilha = []

for row in range(len(Presencas)):
    with open("Candidatos com 20 horas.txt", "a") as arquivo:
        arquivo.write('\n' + str(Presencas[row]))
